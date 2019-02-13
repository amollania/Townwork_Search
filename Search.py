import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
import re
import os
import time


def page_arrow(url, category):
	fileDir = os.path.dirname(os.path.abspath(__file__))

	# Checking file
	if not os.path.exists(fileDir+"/data.xlsx"):
		wb = Workbook()
		ws = wb.active
		wb.save(fileDir+"/data.xlsx")
		print("Making EXCEL")

	# Openpyxl
	book = openpyxl.load_workbook(fileDir+"/data.xlsx")
	sheet = book.worksheets[0]
	ro = 2
	co = 1
	def sheets(nu,name_get):
		sheet.cell(row=ro,column=co+nu).value = name_get

	# Text box
	page = "&page="
	i = 1
	con_box0 = "job-cassette-lst-wrap"
	con_box1 = "job-lst-main-box-inner"
	con_box2 = "job-detail-box-tbl"
	con_box3 = "job-ditail-tbl-inner"

	while True:
		# BeautifulSoup
		req = requests.get(url+category+page+str(i))
		i += 1
		html = req.text
		soup = BeautifulSoup(html, 'html.parser')
		soup = soup.body

		# Stop in Result of 404 Page
		if "error-message" in str(soup):
			print("Not found, Because I don't any more")
			break

		# import re
		p = re.compile('''<a[^>]+href=["'](.*?)["']''')

		# Worktown
		# con_box0
		data = soup.find("div", class_=con_box0)

		# Search http://
		# con_box1
		link_list = data.find_all("a",class_=con_box1)
		print("Catch")

		link = []
		for title in link_list:
			with open(fileDir+'/wheel.txt', 'a', encoding="utf-8") as f:
				find = p.findall(str(title))
				link.append(find[0])
				f.write(url+find.pop()+"\n")
		link = list(set(link))

		# con_box2
		for l in link:
			time.sleep(0.2)
			try:
				con = requests.get(url+l)
				con = BeautifulSoup(con.text, 'html.parser')
				con = con.find_all("div", class_ = con_box2) #Main Content box
				for detail in con:
					if "会社住所" in detail.text:
						print("Search Detail...")
						detail_list = detail.find_all("dl", class_ = con_box3)
						for d in detail_list:
							if "社名" in d.text:
								company_name = d.find("p").text
								sheets(1,company_name)
							elif "会社事業" in d.text:
								company_career = d.find("p").text
								sheets(2,company_career)
							elif "住所" in d.text:
								company_address = d.find("p").text
								sheets(3,company_address)
							elif "ホームページリンク" in d.text:
								company_homapage = d.find("a").text
								sheets(4,company_homapage)
							else:
								print("I don't found Detail")
						ro+=1
						book.save(fileDir+"/data.xlsx")
						print("Catch Company Detail! EXCEL save!")
			except:
				pass
		print("Success")

